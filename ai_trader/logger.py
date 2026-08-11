"""Trade Logger — logs trading signals to the Excel tracker.

Appends a row to the Trade_Log_Tracker.xlsx file when a trade is confirmed.
Uses openpyxl to load the workbook, find the first empty row, and write the values.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

from ai_trader.signal import SignalContext

logger = logging.getLogger(__name__)

# The tracker template pre-builds formulas in rows 3-500 (Capital Risked, P&L,
# P&L %, R Multiple, Status) that the Dashboard reads via ``'Trade Log'!x3:x500``.
# New trades must land in the first empty row within that range: ``ws.append()``
# would write past it (row 501), outside every Dashboard formula, and the trade
# would never show up there.
_TEMPLATE_FIRST_ROW = 3
_TEMPLATE_LAST_ROW = 500

class LoggerError(RuntimeError):
    """Failed to log the trade (e.g. file is open/locked by another process)."""


class TradeLogger:
    """Logs confirmed trading signals to the Excel tracker."""

    def __init__(self, excel_path: str | Path = "Trade_Log_Tracker.xlsx", sheet_name: str = "Trade Log") -> None:
        self.excel_path = Path(excel_path)
        self.sheet_name = sheet_name

    def log_signal(self, ctx: SignalContext) -> None:
        """Append the signal as a new row in the tracker.
        
        Raises:
            LoggerError: if the file cannot be read or written (e.g. open in Excel).
        """
        try:
            import openpyxl
        except ImportError as exc:
            raise LoggerError("openpyxl is not installed (pip install openpyxl)") from exc

        if not self.excel_path.exists():
            raise LoggerError(f"Tracker file not found: {self.excel_path}")

        try:
            wb = openpyxl.load_workbook(self.excel_path)
        except Exception as exc:
            raise LoggerError(f"Failed to open workbook (is it open in Excel?): {exc}") from exc

        if self.sheet_name in wb.sheetnames:
            ws = wb[self.sheet_name]
        else:
            ws = wb.active

        dt = datetime.fromtimestamp(ctx.captured_at)
        signal = ctx.signal

        # Write only the input columns (A-J) plus Notes (T). Columns M (Capital
        # Risked), P (P&L), Q (P&L %), R (R Multiple) and S (Status) hold
        # pre-built formulas in the template; overwriting them would break the
        # tracker, so they are never written here. ``position_size_pct`` has no
        # column yet and is intentionally omitted until one exists.
        values = {
            1: dt.strftime("%Y-%m-%d"),   # Date
            2: dt.strftime("%H:%M:%S"),   # Time
            3: signal.market,             # Market
            4: ctx.symbol,                # Symbol
            5: signal.action.upper(),     # Direction
            6: ctx.model,                 # Signal Source
            7: signal.confidence,         # AI Confidence %
            8: signal.entry,              # Entry Price
            9: signal.stop_loss,          # Stop Loss
            10: signal.target,            # Target
            20: signal.reasoning,         # Notes
        }
        # Quantity (K) and Currency (L) are written only when the signal
        # actually provides them (the schema does not today).
        for field, column in (("quantity", 11), ("currency", 12)):
            real_value = getattr(signal, field, None)
            if real_value is not None:
                values[column] = real_value

        try:
            target_row = self._first_empty_row(ws)
            for column, value in values.items():
                if value is not None:
                    ws.cell(row=target_row, column=column, value=value)
            wb.save(self.excel_path)
        except LoggerError:
            raise
        except Exception as exc:
            raise LoggerError(f"Failed to save workbook (is it open in Excel?): {exc}") from exc

        logger.info("Trade logged to %s for %s (%s)", self.excel_path.name, ctx.symbol, signal.action)

    @staticmethod
    def _first_empty_row(ws) -> int:
        """First data row (from row 3 down) whose Date column is empty.

        Rows 3-500 carry the template's pre-built formulas and the Dashboard
        only reads within that range, so new trades must reuse one of those
        rows rather than append past the range.

        Raises:
            LoggerError: if every template row already contains data.
        """
        last_row = max(ws.max_row, _TEMPLATE_LAST_ROW)
        for row in range(_TEMPLATE_FIRST_ROW, last_row + 1):
            if ws.cell(row=row, column=1).value in (None, ""):
                return row
        raise LoggerError(f"Tracker is full: rows {_TEMPLATE_FIRST_ROW}-{last_row} all contain data")
