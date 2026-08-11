from __future__ import annotations

import shutil
from pathlib import Path
from unittest import mock

import pytest

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from ai_trader.logger import LoggerError, TradeLogger
from ai_trader.signal import SignalContext, TradingSignal


@pytest.fixture
def mock_ctx() -> SignalContext:
    signal = TradingSignal(
        action="buy",
        confidence=85.0,
        entry=150.5,
        stop_loss=148.0,
        target=155.0,
        position_size_pct=2.0,
        timeframe="15m",
        reasoning="Strong momentum",
        market="NSE",
        symbol="RELIANCE"
    )
    return SignalContext(
        symbol="RELIANCE",
        signal=signal,
        captured_at=1690000000.0,  # Specific timestamp for testing
        provider="ollama",
        model="qwen2.5vl:3b"
    )


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
def test_trade_logger_appends_row(tmp_path: Path, mock_ctx: SignalContext) -> None:
    # Create a blank workbook
    excel_path = tmp_path / "Test_Tracker.xlsx"
    wb = openpyxl.Workbook()
    # Rename active sheet to "Trade Log"
    ws = wb.active
    ws.title = "Trade Log"
    # Append headers
    headers = [
        'Date', 'Time', 'Market', 'Symbol', 'Direction', 'Signal Source', 
        'AI Confidence %', 'Entry Price', 'Stop Loss', 'Target', 'Quantity', 'Currency', 
        'Capital Risked', 'Exit Price', 'Exit Time', 'P&L', 'P&L %', 'R Multiple', 'Status', 'Notes'
    ]
    ws.append(headers)
    wb.save(excel_path)

    # Test the logger
    logger = TradeLogger(excel_path=excel_path)
    logger.log_signal(mock_ctx)

    # Verify the contents
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Trade Log"]
    rows = list(ws.rows)
    # Header row 1, row 2 left empty (targeting starts at row 3), data row 3.
    assert len(rows) == 3

    row_data = [cell.value for cell in rows[2]]

    import datetime
    dt = datetime.datetime.fromtimestamp(mock_ctx.captured_at)

    assert row_data[0] == dt.strftime("%Y-%m-%d")  # Date
    assert row_data[1] == dt.strftime("%H:%M:%S")  # Time
    assert row_data[2] == "NSE"  # Market
    assert row_data[3] == "RELIANCE"  # Symbol
    assert row_data[4] == "BUY"  # Direction
    assert row_data[5] == "qwen2.5vl:3b"  # Signal Source
    assert row_data[6] == 85.0  # AI Confidence %
    assert row_data[7] == 150.5  # Entry Price
    assert row_data[8] == 148.0  # Stop Loss
    assert row_data[9] == 155.0  # Target
    # Quantity / Currency are only written when the signal provides them.
    assert row_data[10] is None  # Quantity
    assert row_data[11] is None  # Currency
    # Formula columns are left untouched (blank workbook → empty cells).
    assert row_data[12] is None  # Capital Risked
    assert row_data[18] is None  # Status
    assert row_data[19] == "Strong momentum"  # Notes


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
def test_trade_logger_writes_into_template_row_3(tmp_path: Path, mock_ctx: SignalContext) -> None:
    """Regression (Task 1): a trade must land in the first empty template row
    (row 3), not row 501 past the Dashboard's ``'Trade Log'!x3:x500`` ranges —
    and the pre-built Capital Risked / Status formulas in that row must survive."""
    # Copy the REAL tracker template so the pre-built formula range is present.
    template = Path(__file__).resolve().parent.parent / "Trade_Log_Tracker.xlsx"
    if not template.exists():
        pytest.skip("Trade_Log_Tracker.xlsx template not present in repo root")
    excel_path = tmp_path / "Trade_Log_Tracker.xlsx"
    shutil.copy2(template, excel_path)

    logger = TradeLogger(excel_path=excel_path)
    logger.log_signal(mock_ctx)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Trade Log"]

    import datetime
    dt = datetime.datetime.fromtimestamp(mock_ctx.captured_at)

    # Row 2 is the template's sample row; row 3 is the first empty data row.
    assert ws.cell(row=3, column=1).value == dt.strftime("%Y-%m-%d")  # Date
    assert ws.cell(row=3, column=4).value == "RELIANCE"  # Symbol
    assert ws.cell(row=3, column=20).value == "Strong momentum"  # Notes (T)

    # The old ws.append() bug wrote to row 501, outside the Dashboard ranges.
    assert ws.cell(row=501, column=1).value in (None, "")

    # Formula columns for the written row are preserved, not clobbered.
    m3 = ws.cell(row=3, column=13).value  # Capital Risked
    assert isinstance(m3, str) and m3.startswith("=IF(")
    s3 = ws.cell(row=3, column=19).value  # Status
    assert isinstance(s3, str) and s3.startswith("=IF(")


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
def test_trade_logger_file_not_found(tmp_path: Path, mock_ctx: SignalContext) -> None:
    missing_path = tmp_path / "Missing.xlsx"
    logger = TradeLogger(excel_path=missing_path)
    with pytest.raises(LoggerError, match="Tracker file not found"):
        logger.log_signal(mock_ctx)


def test_trade_logger_no_openpyxl(mock_ctx: SignalContext) -> None:
    # Simulate missing openpyxl
    with mock.patch.dict("sys.modules", {"openpyxl": None}):
        logger = TradeLogger(excel_path="dummy.xlsx")
        with pytest.raises(LoggerError, match="openpyxl is not installed"):
            logger.log_signal(mock_ctx)
